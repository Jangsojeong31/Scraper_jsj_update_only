"""
법제처 엑셀 파일 파서
법규목록.xlsx 파일을 파싱하여 데이터를 추출합니다.

커맨드 라인 사용법:

    1. 약자 키워드로 필터링 (권장):
       python excel_parser.py --keyword krx
       python excel_parser.py -k law
       python excel_parser.py -k kfb

    2. 한글 키워드로 필터링:
       python excel_parser.py --keyword 한국거래소
       python excel_parser.py -k 금융감독원

    3. 키워드 없이 실행 (비어있는 항목만 필터링):
       python excel_parser.py
       # '협회 등' 컬럼이 비어있는 항목만 추출

    4. 다른 컬럼으로 필터링:
       python excel_parser.py --keyword kfb --column 자율규제
       python excel_parser.py -k kfb -c 자율규제

    5. 출력 파일 지정:
       python excel_parser.py --keyword krx --output-list KRX_Scraper/input/list.csv
       python excel_parser.py -k krx --output-list KRX_Scraper/input/list.csv

    6. 파일 구조 확인:
       python excel_parser.py --inspect

사용 가능한 약자 키워드:
    law      법제처 (빈 값으로 필터링)
    bok      한국은행
    crefia   여신금융협회
    fsb      저축은행중앙회
    kfb      은행연합회
    kofia    금융투자협회
    krx      한국거래소

주요 옵션:
    --keyword, -k      필터링할 키워드 (약자 또는 한글 키워드, 지정하지 않으면 '협회 등' 컬럼이 비어있는 항목만 필터링)
    --column, -c       필터링할 컬럼명 (기본값: '협회 등')
    --file, -f         엑셀 파일 경로 (기본값: 프로젝트 루트/법규목록.xlsx)
    --sheet, -s        파싱할 시트 이름 (기본값: 첫 번째 시트)
    --output-list      법령 목록 CSV 출력 파일 경로 (지정하지 않으면 프로젝트 루트/list.csv)
    --output-csv       전체 데이터 CSV 출력 파일 경로
    --output-json     전체 데이터 JSON 출력 파일 경로
    --inspect          파일 구조만 확인하고 파싱하지 않음

Python 코드에서 사용:
    from excel_parser import LawExcelParser
    
    parser = LawExcelParser()
    records = parser.parse_excel(filter_association_keyword='한국거래소')
    parser.save_to_law_list_csv(records, output_path="KRX_Scraper/input/list.csv")
"""
import sys
from pathlib import Path

# 프로젝트 루트를 sys.path에 추가
def find_project_root():
    """common 디렉토리를 찾을 때까지 상위 디렉토리로 이동"""
    try:
        current = Path(__file__).resolve().parent
    except NameError:
        current = Path.cwd()
    
    while current != current.parent:
        if (current / 'common').exists() and (current / 'common' / 'base_scraper.py').exists():
            return current
        current = current.parent
    
    return Path.cwd()

project_root = find_project_root()
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import pandas as pd
import openpyxl
from typing import List, Dict, Optional
import json
import csv
import os


class LawExcelParser:
    """법제처 엑셀 파일 파서"""
    
    def __init__(self, excel_path: str = None):
        """
        Args:
            excel_path: 엑셀 파일 경로 (None이면 프로젝트 루트의 법규목록.xlsx 사용)
        """
        if excel_path is None:
            excel_path = project_root / "법규목록.xlsx"
        
        self.excel_path = Path(excel_path)
        if not self.excel_path.is_absolute():
            self.excel_path = project_root / excel_path
    
    def inspect_file(self) -> Dict:
        """
        엑셀 파일 구조를 확인합니다.
        
        Returns:
            파일 구조 정보 딕셔너리
        """
        if not self.excel_path.exists():
            print(f"⚠ 파일을 찾을 수 없습니다: {self.excel_path}")
            return {}
        
        print(f"✓ 파일 확인: {self.excel_path}")
        
        # openpyxl로 시트 정보 확인
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        sheet_names = wb.sheetnames
        print(f"\n시트 목록: {sheet_names}")
        
        info = {
            'file_path': str(self.excel_path),
            'sheet_names': sheet_names,
            'sheets': {}
        }
        
        # 각 시트 정보 확인
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            print(f"\n=== 시트: {sheet_name} ===")
            print(f"  행 수: {ws.max_row}, 열 수: {ws.max_column}")
            
            # max_row가 None인 경우 처리 (빈 시트)
            max_row = ws.max_row if ws.max_row is not None else 0
            
            # 첫 5행 데이터 확인
            print(f"\n  첫 5행 데이터:")
            if max_row > 0:
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, max_row), values_only=True), 1):
                    print(f"    {i}: {row}")
            else:
                print("    (데이터 없음)")
            
            # 헤더 행 찾기 (첫 번째 비어있지 않은 행)
            header_row = None
            if max_row > 0:
                for row_idx in range(1, min(10, max_row + 1)):
                    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                    if any(cell for cell in row if cell):
                        header_row = row_idx
                        break
            
            headers = []
            if header_row:
                headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
                print(f"\n  헤더 행 ({header_row}행): {headers}")
            
            info['sheets'][sheet_name] = {
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'header_row': header_row,
                'headers': headers
            }
        
        wb.close()
        return info
    
    def parse_excel(self, sheet_name: Optional[str] = None, header_row: Optional[int] = None, 
                    filter_association_keyword: Optional[str] = None, 
                    filter_column: str = '협회 등') -> List[Dict]:
        """
        엑셀 파일을 파싱하여 데이터를 추출합니다.
        
        Args:
            sheet_name: 파싱할 시트 이름 (None이면 첫 번째 시트)
            header_row: 헤더 행 번호 (None이면 자동 감지)
            filter_association_keyword: 필터링할 키워드 (None이면 필터링 안 함, 빈 문자열이면 비어있는 항목만)
            filter_column: 필터링할 컬럼명 (기본값: '협회 등')
            
        Returns:
            추출된 데이터 리스트 (법령명, 구분, 검색 URL 포함)
        """
        if not self.excel_path.exists():
            print(f"⚠ 파일을 찾을 수 없습니다: {self.excel_path}")
            return []
        
        print(f"✓ 엑셀 파일 파싱 시작: {self.excel_path}")
        
        # pandas로 읽기
        try:
            # 모든 시트 읽기
            excel_file = pd.ExcelFile(self.excel_path)
            
            if sheet_name is None:
                sheet_name = excel_file.sheet_names[0]
            
            print(f"  시트: {sheet_name}")
            
            # 헤더 행 자동 감지
            if header_row is None:
                # 첫 10행을 확인하여 헤더 찾기
                for test_row in range(1, min(10, 100)):
                    try:
                        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, 
                                          header=test_row-1, nrows=1)
                        if len(df.columns) > 0 and not df.isna().all().all():
                            header_row = test_row
                            break
                    except:
                        continue
                
                if header_row is None:
                    header_row = 0  # 기본값: 첫 번째 행
            
            print(f"  헤더 행: {header_row}")
            
            # 데이터 읽기
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=header_row-1 if header_row > 0 else 0)
            
            print(f"  읽은 데이터: {len(df)}행, {len(df.columns)}열")
            print(f"  컬럼명: {list(df.columns)}")
            
            # NaN 값 처리
            df = df.fillna('')  # NaN을 빈 문자열로 변환
            
            # 필터링 적용
            if filter_association_keyword is not None and filter_column in df.columns:
                before_count = len(df)
                if filter_association_keyword == '':
                    # 빈 문자열이면 비어있는 항목만 필터링
                    df = df[df[filter_column].astype(str).str.strip() == '']
                    print(f"  '{filter_column}' 필터링 (비어있는 항목만): {before_count}행 → {len(df)}행")
                else:
                    # 특정 키워드가 포함된 항목만 필터링
                    df = df[df[filter_column].astype(str).str.contains(filter_association_keyword, na=False)]
                    print(f"  '{filter_column}' 필터링 (키워드: '{filter_association_keyword}'): {before_count}행 → {len(df)}행")
            
            # 검색 URL 매핑
            search_urls = {
                '법령': 'https://www.law.go.kr/lsSc.do?menuId=1&subMenuId=15&tabMenuId=81&query=',
                '감독규정': 'https://www.law.go.kr/admRulSc.do?menuId=5&subMenuId=41&tabMenuId=183&query=',
            }
            
            # 데이터를 딕셔너리 리스트로 변환 및 법령명 추출
            # '법', '령', '행정규칙', '자율규제' 컬럼에서 값이 있는 모든 항목을 가져옴 (중복 제거 없음)
            records = []
            
            for idx, row in df.iterrows():
                # 구분 값 확인 (비어있어도 됨)
                division = str(row.get('구분', '')).strip()
                
                # 검색 URL 결정 (구분이 없으면 기본값: 법령 URL)
                search_url = search_urls.get(division, search_urls['법령']) if division else search_urls['법령']
                
                # 약칭 추출 (참고용)
                abbreviation = str(row.get('약칭', '')).strip()
                
                # '법', '령', '행정규칙', '자율규제' 컬럼에서 값이 있는 모든 항목을 레코드로 추가
                # 구분이 비어있어도 법령명이 있으면 추출
                # 구분이 비어있으면 빈 문자열로 유지 (기본값 넣지 않음)
                # 법 컬럼 확인
                law_value = str(row.get('법', '')).strip()
                if law_value:
                    record = {
                        '법령명': law_value,
                        '법령유형': '법',
                        '구분': division,  # 구분이 비어있으면 빈 문자열로 유지
                        '검색_URL': search_url,
                        '약칭': abbreviation,
                    }
                    records.append(record)
                
                # 령 컬럼 확인
                decree_value = str(row.get('령', '')).strip()
                if decree_value:
                    record = {
                        '법령명': decree_value,
                        '법령유형': '령',
                        '구분': division,  # 구분이 비어있으면 빈 문자열로 유지
                        '검색_URL': search_url,
                        '약칭': abbreviation,
                    }
                    records.append(record)
                
                # 행정규칙 컬럼 확인
                admin_value = str(row.get('행정규칙', '')).strip()
                if admin_value:
                    record = {
                        '법령명': admin_value,
                        '법령유형': '행정규칙',
                        '구분': division,  # 구분이 비어있으면 빈 문자열로 유지
                        '검색_URL': search_url,
                        '약칭': abbreviation,
                    }
                    records.append(record)
                
                # 자율규제 컬럼 확인
                self_reg_value = str(row.get('자율규제', '')).strip()
                if self_reg_value:
                    record = {
                        '법령명': self_reg_value,
                        '법령유형': '자율규제',
                        '구분': division,  # 구분이 비어있으면 빈 문자열로 유지
                        '검색_URL': search_url,
                        '약칭': abbreviation,
                    }
                    records.append(record)
            
            print(f"✓ 파싱 완료: {len(records)}개 레코드")
            print(f"  구분별 통계:")
            division_counts = {}
            for record in records:
                div = record.get('구분', '')
                division_counts[div] = division_counts.get(div, 0) + 1
            for div, count in division_counts.items():
                print(f"    {div}: {count}개")
            
            return records
            
        except Exception as e:
            print(f"✗ 파싱 실패: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def save_to_csv(self, records: List[Dict], output_path: str = None):
        """
        파싱된 데이터를 CSV 파일로 저장합니다.
        
        Args:
            records: 저장할 데이터 리스트
            output_path: 출력 파일 경로 (None이면 Law_Scraper/output/law_excel_parsed.csv)
        """
        if not records:
            print("⚠ 저장할 데이터가 없습니다.")
            return
        
        if output_path is None:
            output_dir = project_root / "Law_Scraper" / "output"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / "law_excel_parsed.csv"
        else:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # CSV 저장
        if records:
            headers = list(records[0].keys())
            with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(records)
            
            print(f"✓ CSV 저장 완료: {output_path}")
    
    def save_to_json(self, records: List[Dict], output_path: str = None):
        """
        파싱된 데이터를 JSON 파일로 저장합니다.
        
        Args:
            records: 저장할 데이터 리스트
            output_path: 출력 파일 경로 (None이면 Law_Scraper/output/law_excel_parsed.json)
        """
        if not records:
            print("⚠ 저장할 데이터가 없습니다.")
            return
        
        if output_path is None:
            output_dir = project_root / "Law_Scraper" / "output"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / "law_excel_parsed.json"
        else:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # JSON 저장
        output_data = {
            'source_file': str(self.excel_path),
            'total_count': len(records),
            'records': records
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ JSON 저장 완료: {output_path}")
    
    def save_to_law_list_csv(self, records: List[Dict], output_path: str = None):
        """
        파싱된 데이터에서 법령명과 구분을 추출하여 Law_Scraper/input/list.csv 형식으로 저장합니다.
        
        Args:
            records: 저장할 데이터 리스트
            output_path: 출력 파일 경로 (None이면 Law_Scraper/input/list.csv)
        """
        if not records:
            print("⚠ 저장할 데이터가 없습니다.")
            return
        
        if output_path is None:
            # 프로젝트 루트 찾기
            current = Path(__file__).resolve().parent
            while current != current.parent:
                if (current / 'common').exists():
                    break
                current = current.parent
            output_path = current / "Law_Scraper" / "input" / "list.csv"
        else:
            output_path = Path(output_path)
        
        # 출력 디렉토리 생성
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 법령명과 구분 추출
        law_data = []
        for record in records:
            law_name = record.get('법령명', '').strip()
            division = record.get('구분', '').strip()
            if law_name:
                # 구분이 비어있으면 '-'로 표시
                division_display = division if division else '-'
                law_data.append({
                    '법령명': law_name,
                    '구분': division_display
                })
        
        # CSV 저장 (헤더 포함) - 컬럼 순서: 구분, 법령명
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['구분', '법령명'])
            writer.writeheader()
            for item in law_data:
                # 컬럼 순서에 맞게 재정렬
                writer.writerow({
                    '구분': item.get('구분', ''),
                    '법령명': item.get('법령명', '')
                })
        
        print(f"✓ 법령 목록 저장 완료: {output_path} ({len(law_data)}개)")


def main():
    """메인 함수"""
    import argparse
    
    # 디렉토리 약자와 협회명 매핑
    KEYWORD_MAPPING = {
        'law': '',  # 법제처 - 빈 값으로 필터링
        'bok': '한국은행',
        'crefia': '여신금융협회',
        'fsb': '저축은행중앙회',
        'kfb': '은행연합회',
        'kofia': '금융투자협회',
        'krx': '한국거래소',
    }
    
    # 디렉토리 약자와 디렉토리명 매핑
    DIRECTORY_MAPPING = {
        'law': 'Law_Scraper',
        'bok': 'BOK_Scraper',
        'crefia': 'CREFIA_Scraper',
        'fsb': 'FSB_Scraper',
        'kfb': 'KFB_Scraper',
        'kofia': 'KOFIA_Scraper',
        'krx': 'KRX_Scraper',
    }
    
    parser = argparse.ArgumentParser(description='법제처 엑셀 파일 파서')
    parser.add_argument('--file', '-f', type=str, default=None, help='엑셀 파일 경로 (기본값: 프로젝트 루트/법규목록.xlsx)')
    parser.add_argument('--sheet', '-s', type=str, default=None, help='파싱할 시트 이름 (기본값: 첫 번째 시트)')
    parser.add_argument('--header', type=int, default=None, help='헤더 행 번호 (기본값: 자동 감지)')
    parser.add_argument('--keyword', '-k', type=str, default=None, 
                       help=f'필터링할 키워드 (기본값: 빈 문자열 - 비어있는 항목만)\n'
                            f'사용 가능한 약자: {", ".join(KEYWORD_MAPPING.keys())}\n'
                            f'또는 직접 한글 키워드 입력 가능')
    parser.add_argument('--column', '-c', type=str, default='협회 등', help='필터링할 컬럼명 (기본값: 협회 등)')
    parser.add_argument('--inspect', action='store_true', help='파일 구조만 확인하고 파싱하지 않음')
    parser.add_argument('--output-csv', type=str, default=None, help='CSV 출력 파일 경로')
    parser.add_argument('--output-json', type=str, default=None, help='JSON 출력 파일 경로')
    parser.add_argument('--output-list', type=str, default=None, help='법령 목록 CSV 출력 파일 경로')
    
    args = parser.parse_args()
    
    parser_obj = LawExcelParser(excel_path=args.file)
    
    # 파일 구조 확인
    info = parser_obj.inspect_file()
    
    if args.inspect:
        print("\n=== 파일 구조 확인 완료 ===")
        return
    
    # 데이터 파싱
    print("\n=== 데이터 파싱 시작 ===")
    
    # 키워드가 없으면 모든 키워드에 대해 list.csv 생성
    if args.keyword is None:
        for key, mapped_keyword in KEYWORD_MAPPING.items():
            print(f"\n--- '{key}' 키워드 자동 처리 ---")
            records = parser_obj.parse_excel(
                sheet_name=args.sheet,
                header_row=args.header,
                filter_association_keyword=mapped_keyword,
                filter_column=args.column
            )
            
            if not records:
                print(f"  ⚠ '{key}' 결과가 없어 건너뜀")
                continue
            
            if key == 'law':
                directories = ['Law_Scraper', 'Law_LegNotice_Scraper', 'Moleg_Scraper']
                for directory_name in directories:
                    list_path = project_root / directory_name / "input" / "list.csv"
                    parser_obj.save_to_law_list_csv(records, str(list_path))
            elif key in DIRECTORY_MAPPING:
                directory_name = DIRECTORY_MAPPING[key]
                default_list_path = project_root / directory_name / "input" / "list.csv"
                parser_obj.save_to_law_list_csv(records, str(default_list_path))
            else:
                default_list_path = project_root / "list.csv"
                parser_obj.save_to_law_list_csv(records, str(default_list_path))
        
        print("\n=== 모든 키워드 처리 완료 ===")
        return
    
    # keyword 옵션이 있으면 해당 키워드로 필터링
    keyword_lower = args.keyword.lower()
    if keyword_lower in KEYWORD_MAPPING:
        filter_keyword = KEYWORD_MAPPING[keyword_lower]
        print(f"  키워드 매핑: '{args.keyword}' → '{filter_keyword}'")
    else:
        # 매핑에 없으면 입력한 키워드를 그대로 사용
        filter_keyword = args.keyword
    
    records = parser_obj.parse_excel(
        sheet_name=args.sheet,
        header_row=args.header,
        filter_association_keyword=filter_keyword,
        filter_column=args.column
    )
    
    if records:
        # 저장
        if args.output_csv:
            parser_obj.save_to_csv(records, args.output_csv)
        if args.output_json:
            parser_obj.save_to_json(records, args.output_json)
        
        # list.csv 저장
        if args.output_list:
            # output-list 옵션이 지정되면 해당 경로 사용
            parser_obj.save_to_law_list_csv(records, args.output_list)
        elif keyword_lower == 'law':
            # law 키워드인 경우 세 곳에 모두 저장
            directories = ['Law_Scraper', 'Law_LegNotice_Scraper', 'Moleg_Scraper']
            for directory_name in directories:
                list_path = project_root / directory_name / "input" / "list.csv"
                parser_obj.save_to_law_list_csv(records, str(list_path))
        elif keyword_lower and keyword_lower in DIRECTORY_MAPPING:
            # 약자 키워드가 있고 디렉토리 매핑에 있으면 해당 디렉토리의 input/list.csv에 저장
            directory_name = DIRECTORY_MAPPING[keyword_lower]
            default_list_path = project_root / directory_name / "input" / "list.csv"
            parser_obj.save_to_law_list_csv(records, str(default_list_path))
        else:
            # 키워드가 없거나 매핑에 없으면 프로젝트 루트에 list.csv 생성
            default_list_path = project_root / "list.csv"
            parser_obj.save_to_law_list_csv(records, str(default_list_path))
        
        print(f"\n=== 파싱 완료 ===")
        print(f"총 {len(records)}개 레코드 추출")
        print(f"\n첫 3개 레코드 샘플:")
        for i, record in enumerate(records[:3], 1):
            print(f"\n{i}. {record}")
    else:
        print("\n⚠ 추출된 데이터가 없습니다.")


if __name__ == "__main__":
    main()

