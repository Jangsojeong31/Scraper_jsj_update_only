"""
금융감독원 보도자료 목록에서 첨부파일(HWP, PDF 등)을 모두 추출하고,
보도일을 HWP에서만 추출하는 스크립트 (CSV/Excel/JSON 저장)
"""
import requests
from bs4 import BeautifulSoup
import re
import io
import time
import olefile
import zipfile
import xml.etree.ElementTree as ET
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse
import pandas as pd
import json
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️ pdfplumber가 설치되지 않았습니다. PDF 파일 처리가 제한됩니다.")


# -----------------------------------------------------------
# HWPX 파일에서 텍스트 추출 (ZIP 기반 XML)
# -----------------------------------------------------------
def extract_text_from_hwpx_bytes(hwpx_bytes):
    """HWPX 파일 바이트 데이터를 메모리에서 읽어 텍스트 추출 (ZIP 압축된 XML)"""
    try:
        with zipfile.ZipFile(io.BytesIO(hwpx_bytes)) as zip_file:
            text_content = ""
            
            # HWPX 파일 내부 구조에서 텍스트 찾기
            # 일반적으로 Contents/section0.xml, Contents/section1.xml 등에 텍스트가 있음
            possible_paths = [
                'Contents/section0.xml',
                'Contents/section1.xml',
                'section0.xml',
                'section1.xml',
                'body.xml',
                'Contents/body.xml'
            ]
            
            for path in possible_paths:
                try:
                    if path in zip_file.namelist():
                        xml_data = zip_file.read(path)
                        # XML 파싱
                        root = ET.fromstring(xml_data)
                        
                        # XML에서 텍스트 추출
                        texts = []
                        for elem in root.iter():
                            if elem.text:
                                texts.append(elem.text.strip())
                            if elem.tail:
                                texts.append(elem.tail.strip())
                        
                        text = ' '.join([t for t in texts if t])
                        text = re.sub(r'\s+', ' ', text)
                        
                        if len(text.strip()) > 10:
                            text_content = text
                            break
                except Exception:
                    continue
            
            return text_content
            
    except Exception as e:
        print(f"    ⚠️ HWPX 파일 파싱 오류: {e}")
        return ""


# -----------------------------------------------------------
# HWP 파일에서 텍스트 추출 (OLE2 형식)
# -----------------------------------------------------------
def extract_text_from_hwp_bytes(hwp_bytes):
    """HWP 파일 바이트 데이터를 메모리에서 읽어 텍스트 추출 (OLE2 형식)"""
    try:
        with olefile.OleFileIO(io.BytesIO(hwp_bytes)) as ole:
            text_content = ""
            possible_paths = ['PrvText', 'BodyText/Section0', 'Section0', 'DocInfo', 'BodyText']
            for path in possible_paths:
                if ole.exists(path):
                    data = ole.openstream(path).read()
                    try:
                        text = data.decode('utf-16-le', errors='ignore')
                        text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', text)
                        text = re.sub(r'\s+', ' ', text)

                        if len(text.strip()) > 10:
                            text_content = text
                            break
                    except Exception:
                        pass
            return text_content

    except Exception as e:
        # OLE2 형식이 아니면 HWPX일 수 있음
        return ""


# -----------------------------------------------------------
# PDF 파일에서 텍스트 추출
# -----------------------------------------------------------
def extract_text_from_pdf_bytes(pdf_bytes):
    """PDF 파일 바이트 데이터를 메모리에서 읽어 텍스트 추출"""
    if not PDF_AVAILABLE:
        return ""
    
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text_content = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content += page_text + "\n"
            return text_content.strip()
    except Exception as e:
        print(f"    ⚠️ PDF 파일 파싱 오류: {e}")
        return ""


# -----------------------------------------------------------
# 키워드 주변에서 날짜 추출
# -----------------------------------------------------------
def extract_date_near_keyword(text, keyword, context_range=150):
    """특정 키워드 주변에서 날짜를 찾습니다 (표 형식도 고려)"""
    if not text or not keyword:
        return None
    
    # 키워드 앞뒤 지정된 범위 내에서 날짜 찾기
    pattern = re.compile(
        rf'.{{0,{context_range}}}{re.escape(keyword)}.{{0,{context_range}}}',
        re.IGNORECASE | re.DOTALL
    )
    matches = pattern.finditer(text)
    
    date_patterns = [
        r'(\'?\d{2,4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)',
        r'(\'?\d{2,4}\.\s*\d{1,2}\.\s*\d{1,2}\s*\(?[가-힣]*\)?)',
        r'(\'?\d{2,4}-\d{1,2}-\d{1,2})',
        r'(\'?\d{2,4}/\d{1,2}/\d{1,2})',
        r'(\d{8})',  # 8자리 숫자
        r'(\d{10})',  # 10자리 숫자 (예: 25032011)
    ]
    
    for match in matches:
        context = match.group(0)
        # 컨텍스트에서 날짜 패턴 찾기
        for date_pattern in date_patterns:
            date_match = re.search(date_pattern, context)
            if date_match:
                date_str = date_match.group(1).strip()
                
                # 10자리 숫자 형식 처리 (예: 25032011 -> 2025-03-20)
                if len(date_str) == 10 and date_str.isdigit():
                    year = int(date_str[:2])
                    month = int(date_str[2:4])
                    day = int(date_str[4:6])
                    if year >= 50:
                        full_year = 1900 + year
                    else:
                        full_year = 2000 + year
                    date_str = f"{full_year}.{month}.{day}"
                
                # '25 같은 형식을 2025로 변환
                date_str = normalize_year_format(date_str)
                # 년도가 없으면 현재 년도 추가
                date_str = add_year_if_missing(date_str)
                return date_str
    
    # 표 형식 처리: 키워드 다음 줄이나 같은 줄에 날짜가 있을 수 있음
    # 예: "보 도\n2025.3.20" 또는 "보도일\t2025.3.20"
    lines = text.split('\n')
    for i, line in enumerate(lines):
        if keyword in line:
            # 같은 줄에서 찾기
            for date_pattern in date_patterns:
                date_match = re.search(date_pattern, line)
                if date_match:
                    date_str = date_match.group(1).strip()
                    if len(date_str) == 10 and date_str.isdigit():
                        year = int(date_str[:2])
                        month = int(date_str[2:4])
                        day = int(date_str[4:6])
                        if year >= 50:
                            full_year = 1900 + year
                        else:
                            full_year = 2000 + year
                        date_str = f"{full_year}.{month}.{day}"
                    date_str = normalize_year_format(date_str)
                    date_str = add_year_if_missing(date_str)
                    return date_str
            
            # 다음 줄에서 찾기
            if i + 1 < len(lines):
                next_line = lines[i + 1]
                for date_pattern in date_patterns:
                    date_match = re.search(date_pattern, next_line)
                    if date_match:
                        date_str = date_match.group(1).strip()
                        if len(date_str) == 10 and date_str.isdigit():
                            year = int(date_str[:2])
                            month = int(date_str[2:4])
                            day = int(date_str[4:6])
                            if year >= 50:
                                full_year = 1900 + year
                            else:
                                full_year = 2000 + year
                            date_str = f"{full_year}.{month}.{day}"
                        date_str = normalize_year_format(date_str)
                        date_str = add_year_if_missing(date_str)
                        return date_str
    
    return None


# -----------------------------------------------------------
# 텍스트에서 보도일 추출 (보도일 우선, 없으면 배포일)
# -----------------------------------------------------------
def extract_first_date(text):
    """텍스트에서 보도일을 추출합니다.
    1. 보도일 관련 키워드 주변에서 우선 검색
    2. 보도일을 못 찾으면 배포일 관련 키워드 주변에서 검색
    3. "보도가 배포 시" 같은 복합 패턴 처리
    4. 그래도 못 찾으면 전체 텍스트에서 검색"""
    if not text:
        return None

    # 0단계: "보도시점은 배포시", "보도시점: 배포시", "보 도" 같은 복합 패턴 처리
    # 이 경우 "배포시" 다음에 오는 날짜를 찾아야 함
    complex_patterns = [
        # "보 도" 패턴 (띄어쓰기 포함) - 문서 상단에 자주 나타나는 형식
        r'보\s+도\s*[:：]?\s*(\'?\d{2,4}\.\s*\d{1,2}\.\s*\d{1,2})',
        r'보\s+도\s*[:：]?\s*(\'?\d{2,4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)',
        r'보\s+도\s*[:：]?\s*(\'?\d{2,4}-\d{1,2}-\d{1,2})',
        r'보\s+도\s*[:：]?\s*(\'?\d{2,4}/\d{1,2}/\d{1,2})',
        r'보\s+도\s*[:：]?\s*(\d{10})',  # 10자리 숫자 (예: 25032011)
        # "보도시점은 배포시" 또는 "보도시점: 배포시" 패턴
        r'보도\s*시점\s*[은는:：]\s*배포\s*시\s*[:：]?\s*(\'?\d{2,4}\.\s*\d{1,2}\.\s*\d{1,2})',
        r'보도\s*시점\s*[은는:：]\s*배포\s*시\s*[:：]?\s*(\'?\d{2,4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)',
        r'보도\s*시점\s*[은는:：]\s*배포\s*시\s*[:：]?\s*(\'?\d{2,4}-\d{1,2}-\d{1,2})',
        r'보도\s*시점\s*[은는:：]\s*배포\s*시\s*[:：]?\s*(\'?\d{2,4}/\d{1,2}/\d{1,2})',
        r'보도\s*시점\s*[은는:：]\s*배포\s*시\s*[:：]?\s*(\d{10})',  # 10자리 숫자
        # "보도가 배포 시" 패턴
        r'보도\s*[가와]\s*배포\s*시\s*[:：]?\s*(\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)',
        r'보도\s*[가와]\s*배포\s*시\s*[:：]?\s*(\d{4}\.\s*\d{1,2}\.\s*\d{1,2})',
        r'보도\s*[가와]\s*배포\s*시\s*[:：]?\s*(\d{4}-\d{1,2}-\d{1,2})',
        r'보도\s*[가와]\s*배포\s*시\s*[:：]?\s*(\d{4}/\d{1,2}/\d{1,2})',
        r'보도\s*[가와]\s*배포\s*시\s*[:：]?\s*(\d{10})',  # 10자리 숫자
    ]
    for pattern in complex_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            date_str = match.group(1).strip()
            
            # 10자리 숫자 형식 처리 (예: 25032011 -> 2025-03-20)
            if len(date_str) == 10 and date_str.isdigit():
                year = int(date_str[:2])
                month = int(date_str[2:4])
                day = int(date_str[4:6])
                if year >= 50:
                    full_year = 1900 + year
                else:
                    full_year = 2000 + year
                date_str = f"{full_year}.{month}.{day}"
            
            # '25 같은 형식을 2025로 변환
            date_str = normalize_year_format(date_str)
            date_str = add_year_if_missing(date_str)
            print(f"      ℹ️ 복합 패턴에서 날짜 발견: {date_str}")
            return date_str

    # 1단계: 보도일 관련 키워드 우선 검색 (띄어쓰기 변형 포함)
    # "보도시점" 키워드가 있지만 "배포시"로 연결된 경우는 이미 0단계에서 처리됨
    # "보 도" (띄어쓰기) 패턴도 포함
    press_keywords = [
        '보도시점', '보도 시점', '보도시 점', 
        '보도일', '보도 일', 
        '보도 시', '보도가', 
        '보 도', '보  도',  # 띄어쓰기 변형
        '보도'
    ]
    
    for keyword in press_keywords:
        date = extract_date_near_keyword(text, keyword)
        if date:
            return date
    
    # 2단계: 보도일을 못 찾았으면 배포일 관련 키워드 검색 (배포시 포함)
    release_keywords = ['배포시', '배포 시', '배포일', '배포 일', '배포가', '배포']
    
    for keyword in release_keywords:
        date = extract_date_near_keyword(text, keyword)
        if date:
            print(f"      ℹ️ 보도일을 찾지 못해 배포일을 사용: {date}")
            return date
    
    # 3단계: 키워드 주변에서 못 찾으면 전체 텍스트에서 찾기
    # 3단계: 키워드 주변에서 못 찾으면 전체 텍스트에서 찾기 (2자리 년도 포함)
    date_patterns = [
        r'(\'?\d{2,4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)',
        r'(\'?\d{2,4}\.\s*\d{1,2}\.\s*\d{1,2}\s*\(?[가-힣]*\)?)',
        r'(\'?\d{2,4}-\d{1,2}-\d{1,2})',
        r'(\'?\d{2,4}/\d{1,2}/\d{1,2})',
        r'(\d{8})',  # 8자리 숫자 (예: 20250320)
        r'(\d{10})',  # 10자리 숫자 (예: 25032011 -> 2025-03-20으로 변환)
    ]
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            date_str = match.group(1).strip()
            
            # 10자리 숫자 형식 처리 (예: 25032011 -> 2025-03-20)
            if len(date_str) == 10 and date_str.isdigit():
                year = int(date_str[:2])
                month = int(date_str[2:4])
                day = int(date_str[4:6])
                # 시간 부분은 무시
                if year >= 50:
                    full_year = 1900 + year
                else:
                    full_year = 2000 + year
                date_str = f"{full_year}.{month}.{day}"
            
            # '25 같은 형식을 2025로 변환
            date_str = normalize_year_format(date_str)
            # 년도가 없으면 현재 년도 추가
            date_str = add_year_if_missing(date_str)
            return date_str

    return None


# -----------------------------------------------------------
# 텍스트에서 보도시점 추출
# -----------------------------------------------------------
def extract_press_time(text):
    """텍스트에서 보도시점을 추출합니다"""
    if not text:
        return None
    
    # 보도시점 키워드 주변에서 날짜 찾기
    press_time_keywords = ['보도시점', '보도 시점', '보도시 점']
    
    for keyword in press_time_keywords:
        date = extract_date_near_keyword(text, keyword)
        if date:
            return date
    
    return None


# -----------------------------------------------------------
# 년도 형식 정규화 ('25 -> 2025)
# -----------------------------------------------------------
def normalize_year_format(date_str):
    """'25.9.3 같은 2자리 년도를 2025.9.3로 변환"""
    if not date_str:
        return date_str
    
    # '25 또는 25로 시작하는 패턴 찾기
    # '25.9.3 -> 2025.9.3
    # 25.9.3 -> 2025.9.3 (2자리 년도인 경우)
    pattern = r'^(\'?)(\d{2})(\.\s*\d{1,2}\.\s*\d{1,2})'
    match = re.match(pattern, date_str)
    if match:
        prefix = match.group(1)  # ' 또는 빈 문자열
        year = int(match.group(2))  # 25
        rest = match.group(3)  # .9.3
        
        # 2자리 년도를 4자리로 변환 (50 이상이면 1900년대, 미만이면 2000년대)
        if year >= 50:
            full_year = 1900 + year
        else:
            full_year = 2000 + year
        
        return f"{full_year}{rest}"
    
    # '25년 9월 3일 형식
    pattern = r'^(\'?)(\d{2})\s*년'
    match = re.match(pattern, date_str)
    if match:
        prefix = match.group(1)
        year = int(match.group(2))
        
        if year >= 50:
            full_year = 1900 + year
        else:
            full_year = 2000 + year
        
        return date_str.replace(f"{prefix}{year}년", f"{full_year}년", 1)
    
    return date_str


# -----------------------------------------------------------
# 날짜에 년도가 없으면 현재 년도 추가
# -----------------------------------------------------------
def add_year_if_missing(date_str):
    """날짜 문자열에 년도가 없으면 현재 년도를 추가합니다"""
    if not date_str:
        return date_str
    
    current_year = datetime.now().year
    
    # 년도가 있는지 확인 (4자리 숫자로 시작)
    if re.match(r'^\d{4}', date_str):
        return date_str
    
    # 년도가 없는 경우: "1월 1일", "1.1", "1-1", "1/1" 등
    # 현재 년도를 앞에 추가
    if re.match(r'^\d{1,2}\s*월\s*\d{1,2}\s*일', date_str):
        return f"{current_year}년 {date_str}"
    elif re.match(r'^\d{1,2}\.\s*\d{1,2}', date_str):
        return f"{current_year}.{date_str}"
    elif re.match(r'^\d{1,2}-\d{1,2}', date_str):
        return f"{current_year}-{date_str}"
    elif re.match(r'^\d{1,2}/\d{1,2}', date_str):
        return f"{current_year}/{date_str}"
    elif re.match(r'^\d{4}$', date_str):  # 4자리 숫자만 있는 경우 (월일)
        return f"{current_year}{date_str}"
    
    return date_str


# -----------------------------------------------------------
# 날짜 문자열을 datetime 객체로 변환
# -----------------------------------------------------------
def parse_date_string(date_str):
    """다양한 형식의 날짜 문자열을 datetime 객체로 변환"""
    if not date_str:
        return None
    
    date_str = date_str.strip()
    
    # "2024년 1월 1일" 형식
    match = re.match(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except:
            pass
    
    # "2024. 1. 1" 또는 "2024.1.1" 형식
    match = re.match(r'(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except:
            pass
    
    # "2024-01-01" 형식
    match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except:
            pass
    
    # "2024/01/01" 형식
    match = re.match(r'(\d{4})/(\d{1,2})/(\d{1,2})', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except:
            pass
    
    # "20240101" 형식
    match = re.match(r'(\d{4})(\d{2})(\d{2})', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except:
            pass
    
    return None


# -----------------------------------------------------------
# 날짜가 2025년 이후인지 확인
# -----------------------------------------------------------
def is_after_2025(date_str):
    """날짜 문자열이 2025년 이후인지 확인"""
    if not date_str:
        return True  # 날짜가 없으면 일단 포함 (나중에 필터링)
    
    date_obj = parse_date_string(date_str)
    if not date_obj:
        return True  # 파싱 실패 시 일단 포함
    
    cutoff_date = datetime(2025, 1, 1)
    return date_obj >= cutoff_date


# -----------------------------------------------------------
# 날짜가 지정된 날짜 이후인지 확인
# -----------------------------------------------------------
def is_after_date(date_str, cutoff_date):
    """날짜 문자열이 지정된 날짜 이후인지 확인"""
    if not date_str:
        return True  # 날짜가 없으면 일단 포함 (나중에 필터링)
    
    date_obj = parse_date_string(date_str)
    if not date_obj:
        return True  # 파싱 실패 시 일단 포함
    
    return date_obj >= cutoff_date


# -----------------------------------------------------------
# 다음 페이지가 있는지 확인
# -----------------------------------------------------------
def has_next_page(soup, current_page):
    """현재 페이지에서 다음 페이지가 있는지 확인"""
    try:
        # 다음 페이지 버튼 찾기 (텍스트로 찾기)
        next_texts = ['다음', '>', '▶', 'next', 'Next']
        all_links = soup.find_all('a', href=True)
        for link in all_links:
            text = link.get_text(strip=True)
            if any(next_text in text for next_text in next_texts):
                href = link.get("href", "").strip()
                if href and "pageIndex=" in href:
                    return True
        
        # .next 클래스 찾기
        next_links = soup.select(".next, .paging .next, .pagination .next")
        if next_links:
            return True
        
        # 페이지 번호 링크에서 현재 페이지보다 큰 번호 찾기
        pagination_selectors = [
            "div.paging",
            "div.pagination",
            "div.pageArea",
            ".paging",
            ".pagination",
            ".pageArea",
        ]
        
        for selector in pagination_selectors:
            pagination = soup.select_one(selector)
            if pagination:
                page_links = pagination.select("a[href]")
                for link in page_links:
                    text = link.get_text(strip=True)
                    if text.isdigit() and int(text) > current_page:
                        return True
                    
                    href = link.get("href", "").strip()
                    if href and "pageIndex=" in href:
                        match = re.search(r'pageIndex=(\d+)', href)
                        if match:
                            page_num = int(match.group(1))
                            if page_num > current_page:
                                return True
        
        # 전체 링크에서 다음 페이지 찾기
        for link in all_links:
            href = link.get("href", "").strip()
            if href and "pageIndex=" in href:
                match = re.search(r'pageIndex=(\d+)', href)
                if match:
                    page_num = int(match.group(1))
                    if page_num > current_page:
                        return True
        
        return False
        
    except Exception as e:
        print(f"    ⚠️ 다음 페이지 확인 실패: {e}")
        return False


# -----------------------------------------------------------
# 단일 페이지에서 보도자료 추출
# -----------------------------------------------------------
def scrape_single_page(session, page_url, page_num, total_pages, start_idx=1, cutoff_date=None):
    """단일 페이지에서 보도자료 데이터를 추출합니다
    
    Args:
        cutoff_date: 이 날짜 이후의 보도자료만 수집 (None이면 2025-01-01 사용)
    """
    results = []
    has_recent_data = False  # 페이지 내에 조건에 맞는 데이터가 있는지 여부
    missing_dates_count = 0  # 보도일이 없는 항목 개수
    
    if cutoff_date is None:
        cutoff_date = datetime(2025, 1, 1)
    
    try:
        response = session.get(page_url, timeout=30)
        response.raise_for_status()
        response.encoding = 'utf-8'

        soup = BeautifulSoup(response.text, 'html.parser')
        table = soup.find('table', class_='board_list') or soup.find('table')

        if not table:
            print(f"    ⚠️ 페이지 {page_num}: 테이블을 찾을 수 없습니다.")
            return results, False, False  # (results, should_stop, has_recent_data)

        rows = table.find_all('tr')[1:]
        if not rows:
            print(f"    ⚠️ 페이지 {page_num}: 데이터가 없습니다.")
            return results, False, False

        print(f"\n📄 페이지 {page_num} 처리 중... ({len(rows)}개 항목)")

        # 각 보도자료 행 반복 처리
        for row_idx, row in enumerate(rows, start=start_idx):
            title_link = row.find('a', href=re.compile(r'view\.do'))
            if not title_link:
                continue

            # 제목, 상세 URL
            title = title_link.get_text(strip=True)
            detail_url = urljoin(page_url, title_link['href'])

            # 담당부서
            tds = row.find_all('td')
            department = tds[2].get_text(strip=True) if len(tds) >= 3 else None

            # 첨부파일 (.hwp, .pdf, 등) - 별첨파일 제외
            file_links = []
            attach_links = row.find_all('a', href=re.compile(r'fileDown\.do'))

            for link in attach_links:
                href = urljoin(page_url, link['href'])
                file_name = link.get_text(strip=True)
                
                # 별첨파일 제외
                if '별첨' in file_name or '별 첨' in file_name:
                    continue

                file_links.append({
                    '첨부파일명': file_name,
                    '첨부파일 url': href
                })

            print(f"  [{row_idx}] {title}")
            if not file_links:
                print("      ⚠️ 첨부파일 없음")

            # 신년사 제외
            if '신년사' in title:
                print(f"      ⏹️ 신년사 항목은 제외합니다.")
                continue
            
            # 상세 본문 가져오기 및 등록일 추출
            registration_date = None
            try:
                detail_response = session.get(detail_url, timeout=30)
                detail_response.raise_for_status()
                detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
                content_div = detail_soup.find('div', class_='dbdata')
                content = content_div.get_text(separator='\n', strip=True) if content_div else ''
                content = re.sub(r'\n+', '\n', content.strip())
                
                # 등록일 추출 (보도자료가 아닌 경우 사용)
                # 등록일은 보통 상세페이지의 메타 정보에 있음
                # 여러 패턴 시도
                reg_patterns = [
                    r'등록일[:\s]*(\d{4}[\.\-/]\d{1,2}[\.\-/]\d{1,2})',
                    r'작성일[:\s]*(\d{4}[\.\-/]\d{1,2}[\.\-/]\d{1,2})',
                    r'등록[:\s]*(\d{4}[\.\-/]\d{1,2}[\.\-/]\d{1,2})',
                    r'(\d{4}[\.\-/]\d{1,2}[\.\-/]\d{1,2})',  # 일반적인 날짜 패턴
                ]
                
                # 상세페이지 전체 텍스트에서 등록일 찾기
                page_text = detail_soup.get_text()
                for pattern in reg_patterns:
                    date_match = re.search(pattern, page_text)
                    if date_match:
                        registration_date = date_match.group(1)
                        # 유효한 날짜인지 확인 (2025년 이후)
                        date_obj = parse_date_string(registration_date)
                        if date_obj and date_obj >= datetime(2025, 1, 1):
                            break
                        elif date_obj:
                            registration_date = None  # 2025년 이전이면 무시
                
            except Exception as e:
                print(f"      ⚠️ 상세페이지 크롤링 실패: {e}")
                content = ''

            # 보도자료 파일에서 보도일 추출 (별첨파일 제외)
            date = None
            text_preview = None
            full_text = None  # 전체 텍스트 (보도시점 추출용)

            # 별첨파일 제외하고 보도자료 파일만 필터링
            press_files = [
                f for f in file_links 
                if '별첨' not in f['첨부파일명'] and '별 첨' not in f['첨부파일명']
            ]
            
            # HWP 파일 먼저 시도
            hwp_files = [
                f for f in press_files 
                if f['첨부파일명'].lower().endswith('.hwp')
            ]
            
            for f in hwp_files:
                try:
                    print(f"      📂 HWP 다운로드 중: {f['첨부파일명']}")
                    file_response = session.get(f['첨부파일 url'], timeout=30)
                    file_response.raise_for_status()

                    text = extract_text_from_hwp_bytes(file_response.content)
                    if text:
                        if not full_text:
                            full_text = text
                        if not text_preview:
                            text_preview = text[:200]
                        found_date = extract_first_date(text)
                        print(f"      📅 보도일: {found_date or '추출 실패'}")
                        
                        if found_date:
                            # 날짜가 기준일 이후인지 확인
                            if is_after_date(found_date, cutoff_date):
                                date = found_date
                                print(f"      ✅ 기준일 이후 보도일 확인: {date}")
                                break  # 기준일 이후 날짜를 찾았으면 중단
                            else:
                                print(f"      ⚠️ 보도일이 기준일 이전입니다. 다른 파일을 시도합니다.")
                                # 기준일 이전이면 다른 파일을 시도하기 위해 continue

                except Exception as e:
                    print(f"      ⚠️ HWP 처리 실패 ({f['첨부파일명']}): {e}")
            
            # HWP에서 1년 이내 날짜를 못 찾았으면 HWPX 파일 시도
            if not date:
                hwpx_files = [
                    f for f in press_files 
                    if f['첨부파일명'].lower().endswith('.hwpx')
                ]
                
                for f in hwpx_files:
                    try:
                        print(f"      📂 HWPX 다운로드 중: {f['첨부파일명']}")
                        file_response = session.get(f['첨부파일 url'], timeout=30)
                        file_response.raise_for_status()

                        text = extract_text_from_hwpx_bytes(file_response.content)
                        if text:
                            if not full_text:
                                full_text = text
                            if not text_preview:
                                text_preview = text[:200]
                            found_date = extract_first_date(text)
                            print(f"      📅 보도일 (HWPX): {found_date or '추출 실패'}")
                            
                            if found_date:
                                # 날짜가 기준일 이후인지 확인
                                if is_after_date(found_date, cutoff_date):
                                    date = found_date
                                    print(f"      ✅ 기준일 이후 보도일 확인: {date}")
                                    break  # 기준일 이후 날짜를 찾았으면 중단
                                else:
                                    print(f"      ⚠️ 보도일이 기준일 이전입니다. 다른 파일을 시도합니다.")
                                    # 기준일 이전이면 다른 파일을 시도하기 위해 continue

                    except Exception as e:
                        print(f"      ⚠️ HWPX 처리 실패 ({f['첨부파일명']}): {e}")
            
            # HWP/HWPX에서 1년 이내 날짜를 못 찾았으면 PDF 파일 시도 (별첨 제외)
            if not date:
                pdf_files = [
                    f for f in press_files 
                    if f['첨부파일명'].lower().endswith('.pdf')
                ]
                
                for f in pdf_files:
                    try:
                        print(f"      📂 PDF 다운로드 중: {f['첨부파일명']}")
                        file_response = session.get(f['첨부파일 url'], timeout=30)
                        file_response.raise_for_status()

                        text = extract_text_from_pdf_bytes(file_response.content)
                        if text:
                            if not full_text:
                                full_text = text
                            if not text_preview:
                                text_preview = text[:200]
                            found_date = extract_first_date(text)
                            print(f"      📅 보도일 (PDF): {found_date or '추출 실패'}")
                            
                            if found_date:
                                # 날짜가 기준일 이후인지 확인
                                if is_after_date(found_date, cutoff_date):
                                    date = found_date
                                    print(f"      ✅ 기준일 이후 보도일 확인: {date}")
                                    break  # 기준일 이후 날짜를 찾았으면 중단
                                else:
                                    print(f"      ⚠️ 보도일이 기준일 이전입니다. 다른 파일을 시도합니다.")
                                    # 기준일 이전이면 다른 파일을 시도하기 위해 continue

                    except Exception as e:
                        print(f"      ⚠️ PDF 처리 실패 ({f['첨부파일명']}): {e}")

            # 보도시점 추출 (전체 텍스트 사용)
            press_time = None
            if full_text:
                press_time = extract_press_time(full_text)
            elif text_preview:
                press_time = extract_press_time(text_preview)
            
            # 보도일이 없고 보도시점이 있으면 보도시점을 보도일로 사용
            if not date and press_time:
                date = press_time
                print(f"      ℹ️ 보도일이 없어 보도시점을 보도일로 사용: {date}")
            
            # 보도자료가 아닌 경우(보도참고, 당부사항 등) 등록일을 보도일로 사용
            is_press_release = '보도자료' in title or '보도' in title
            if not date and registration_date and not is_press_release:
                date = registration_date
                print(f"      ℹ️ 보도자료가 아니어서 등록일을 보도일로 사용: {date}")
            
            # 보도일 기준으로만 판단 (등록일은 참고용으로만 사용)
            # 보도일이 있고 기준일 이전이면 스크랩 중단
            if date and not is_after_date(date, cutoff_date):
                print(f"      ⏹️ 보도일({date})이 기준일({cutoff_date.strftime('%Y-%m-%d')}) 이전입니다. 스크랩을 중단합니다.")
                return results, True, False, missing_dates_count  # (results, should_stop, has_recent_data, missing_dates_count)
            
            # 보도일 기준으로 포함 여부 결정
            should_include = False
            if date and is_after_date(date, cutoff_date):
                should_include = True
            elif not date:
                # 보도일이 없어도 일단 포함 (나중에 필터링)
                should_include = True
            
            # 결과 저장 (보도시점과 등록일은 참고용으로만 사용, 저장하지 않음)
            if should_include:
                results.append({
                    '번호': row_idx,
                    '제목': title,
                    '담당부서': department,
                    '보도일': date,
                    '첨부파일': file_links,
                    '첨부파일내용 미리보기': text_preview,
                    '상세페이지URL': detail_url,
                    '내용': content
                })
                # 보도일이 없으면 카운트 증가
                if not date:
                    missing_dates_count += 1
                # 기준일 이후 데이터가 있으면 표시 (보도일 기준)
                if date and is_after_date(date, cutoff_date):
                    has_recent_data = True

            time.sleep(0.5)

    except Exception as e:
        print(f"    ❌ 페이지 {page_num} 처리 오류: {e}")

    # 페이지 내에 기준일 이후 데이터가 없으면 중단 신호 반환
    should_stop = not has_recent_data and len(results) == 0
    
    return results, should_stop, has_recent_data, missing_dates_count  # (results, should_stop, has_recent_data, missing_dates_count)


# -----------------------------------------------------------
# 기존 데이터 로드
# -----------------------------------------------------------
def load_existing_data(json_file="results.json"):
    """기존에 저장된 데이터를 로드합니다"""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print(f"📂 기존 데이터 로드: {len(data)}건 발견")
            return data
    except FileNotFoundError:
        print("📂 기존 데이터 없음 (처음부터 시작)")
        return []
    except Exception as e:
        print(f"⚠️ 기존 데이터 로드 실패: {e} (처음부터 시작)")
        return []


# -----------------------------------------------------------
# 보도자료 목록 스크래핑 (모든 페이지)
# -----------------------------------------------------------
def scrape_press_releases(base_url, total_pages=2010, resume=True):
    session = requests.Session()
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})

    results = []
    
    # 기준일 설정: 2025년 1월 1일
    default_cutoff_date = datetime(2025, 1, 1)
    cutoff_date = default_cutoff_date
    
    try:
        # 기존 데이터 로드 및 기준일 설정
        all_results = []
        item_counter = 1
        start_page = 1
        
        if resume:
            existing_data = load_existing_data("results.json")
            if existing_data:
                all_results = existing_data
                item_counter = len(existing_data) + 1
                
                # 기존 데이터에서 가장 최신 보도일 찾기
                latest_date = None
                for item in existing_data:
                    date_str = item.get('보도일')
                    if date_str:
                        date_obj = parse_date_string(date_str)
                        if date_obj:
                            if latest_date is None or date_obj > latest_date:
                                latest_date = date_obj
                
                if latest_date:
                    # 가장 최신 보도일 이후의 데이터만 수집
                    cutoff_date = latest_date
                    print(f"📅 기존 데이터에서 가장 최신 보도일: {latest_date.strftime('%Y-%m-%d')}")
                    print(f"📅 이 날짜 이후의 신규 보도자료만 수집합니다.")
                else:
                    # 보도일이 없으면 기본 기준일 사용
                    cutoff_date = default_cutoff_date
                    print(f"📅 기존 데이터에 보도일이 없어 기본 기준일({default_cutoff_date.strftime('%Y-%m-%d')})을 사용합니다.")
                
                print(f"🔄 이어서 진행: {len(existing_data)}건 저장됨, 신규 데이터 추가 중...\n")
                import sys
                sys.stdout.flush()
            else:
                print(f"📅 최초 실행: {default_cutoff_date.strftime('%Y-%m-%d')} 이후 보도자료만 수집합니다.\n")
        else:
            print(f"📅 최초 실행: {default_cutoff_date.strftime('%Y-%m-%d')} 이후 보도자료만 수집합니다.\n")
        
        print("📢 보도자료 목록 처리 중...")
        print("=" * 70)
        print(f"📅 수집 기준일: {cutoff_date.strftime('%Y-%m-%d')} 이후")
        print("=" * 70)
        
        # URL에서 기본 파라미터 추출
        parsed = urlparse(base_url)
        params = parse_qs(parsed.query)
        
        page_num = start_page
        save_interval = 10  # 10개씩 중간 저장
        
        start_time = time.time()
        total_missing_dates = 0  # 전체 누락된 보도일 개수
        
        # 시간 포맷팅 함수
        def format_time(seconds):
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            secs = int(seconds % 60)
            if hours > 0:
                return f"{hours}시간 {minutes}분 {secs}초"
            elif minutes > 0:
                return f"{minutes}분 {secs}초"
            else:
                return f"{secs}초"
        
        while True:  # 1년 이내 데이터가 없을 때까지 계속
            # 페이지 URL 생성
            params['pageIndex'] = [str(page_num)]
            new_query = urlencode(params, doseq=True)
            page_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))
            
            # 페이지 가져오기
            try:
                response = session.get(page_url, timeout=30)
                response.raise_for_status()
                response.encoding = 'utf-8'
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # 테이블 확인
                table = soup.find('table', class_='board_list') or soup.find('table')
                if not table:
                    print(f"  ⚠️ 페이지 {page_num}: 테이블을 찾을 수 없습니다. 종료합니다.")
                    break
                
                rows = table.find_all('tr')[1:]
                if not rows:
                    print(f"  ⚠️ 페이지 {page_num}: 데이터가 없습니다. 종료합니다.")
                    break
                
                # 단일 페이지 스크래핑
                page_results, should_stop, has_recent_data, missing_dates = scrape_single_page(session, page_url, page_num, 0, start_idx=item_counter, cutoff_date=cutoff_date)
                
                if should_stop:
                    print(f"  ⏹️ 페이지 {page_num}에 기준일 이후 데이터가 없습니다. 스크랩을 중단합니다.")
                    break
                
                if not page_results:
                    print(f"  ⚠️ 페이지 {page_num}: 추출된 데이터가 없습니다. 종료합니다.")
                    break
                
                # 기존 데이터와 중복 확인 (상세페이지URL 기준)
                existing_urls = {item.get('상세페이지URL') for item in all_results}
                new_results = []
                for result in page_results:
                    # 중복이 아니고 기준일 이후인 경우만 추가
                    if result.get('상세페이지URL') not in existing_urls:
                        result['번호'] = item_counter
                        item_counter += 1
                        new_results.append(result)
                
                all_results.extend(new_results)
                total_missing_dates += missing_dates
                
                if len(new_results) < len(page_results):
                    print(f"  ℹ️ 중복 항목 {len(page_results) - len(new_results)}개 제외됨")
                
                # 진행률 계산
                elapsed_time = time.time() - start_time
                avg_time_per_item = elapsed_time / len(all_results) if len(all_results) > 0 else 0
                
                # 보도일 추출 성공률 계산
                total_with_dates = len(all_results) - total_missing_dates
                success_rate = (total_with_dates / len(all_results) * 100) if len(all_results) > 0 else 0
                
                print(f"  ✅ 페이지 {page_num} 완료 | "
                      f"추출: {len(page_results)}개 | 신규: {len(new_results)}개 | 누적: {len(all_results)}개 | "
                      f"보도일 누락: {total_missing_dates}개 ({100-success_rate:.1f}%) | "
                      f"경과: {format_time(elapsed_time)} | "
                      f"평균: {avg_time_per_item:.1f}초/건")
                import sys
                sys.stdout.flush()  # 실시간 출력을 위해 버퍼 플러시
                
                # 중간 저장 (10개씩)
                if len(all_results) % save_interval == 0:
                    print(f"\n  💾 중간 저장 중... (누적 {len(all_results)}개)")
                    save_results(all_results, 
                               csv_file="results.csv", 
                               excel_file="results.xlsx", 
                               json_file="results.json")
                    print(f"  ✅ 중간 저장 완료\n")
                
                # 다음 페이지 확인 (안전장치)
                if not has_next_page(soup, page_num):
                    print(f"\n  ⚠️ 다음 페이지를 찾을 수 없습니다.")
                    print(f"  계속 진행합니다...\n")
                
                page_num += 1
                
                # 페이지 간 대기
                time.sleep(1)
                
            except Exception as e:
                print(f"  ❌ 페이지 {page_num} 처리 오류: {e}")
                # 오류 발생 시 중간 저장
                if all_results:
                    print(f"\n  💾 오류 발생으로 중간 저장 중...")
                    save_results(all_results, 
                               csv_file="results.csv", 
                               excel_file="results.xlsx", 
                               json_file="results.json")
                # 연속 오류가 발생하면 종료
                break
        
        results = all_results
        total_time = time.time() - start_time
        
        # 최종 통계
        total_with_dates = len(all_results) - total_missing_dates
        success_rate = (total_with_dates / len(all_results) * 100) if len(all_results) > 0 else 0
        
        print(f"\n📊 스크랩 완료")
        print(f"  - 처리 페이지: {page_num}페이지")
        print(f"  - 총 수집 데이터: {len(all_results)}개")
        print(f"  - 보도일 추출 성공: {total_with_dates}개 ({success_rate:.1f}%)")
        print(f"  - 보도일 누락: {total_missing_dates}개 ({100-success_rate:.1f}%)")
        print(f"  - 소요 시간: {format_time(total_time)}")
        if len(all_results) > 0:
            print(f"  - 평균 처리 시간: {total_time/len(all_results):.1f}초/건")

    except Exception as e:
        print(f"❌ 처리 오류: {e}")
        # 오류 발생 시에도 중간 저장
        if 'all_results' in locals() and all_results:
            print(f"\n  💾 오류 발생으로 중간 저장 중...")
            save_results(all_results, 
                       csv_file="results.csv", 
                       excel_file="results.xlsx", 
                       json_file="results.json")

    return results


# -----------------------------------------------------------
# CSV / Excel / JSON 저장 함수
# -----------------------------------------------------------
def save_results(results, csv_file="results.csv", excel_file="results.xlsx", json_file="results.json"):
    if not results:
        print("❌ 저장할 결과가 없습니다.")
        return

    df = pd.DataFrame(results)

    # 첨부파일 리스트 → 문자열 변환
    df['첨부파일'] = df['첨부파일'].apply(
        lambda lst: ', '.join([f"{f['첨부파일명']} ({f['첨부파일 url']})" for f in lst]) if lst else ''
    )

    df.fillna('', inplace=True)

    # 컬럼 순서 정렬 (보도시점과 등록일은 저장하지 않음)
    columns = ['번호', '제목', '보도일', '상세페이지URL', '첨부파일', '담당부서', '내용']
    # 존재하는 컬럼만 선택
    available_columns = [col for col in columns if col in df.columns]
    df = df[available_columns]

    # CSV 저장
    df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    print(f"📄 CSV 저장 완료: {csv_file}")

    # Excel 저장
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='보도자료')
        ws = writer.sheets['보도자료']

        # 열 너비 자동 조정
        for i, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_len, 50)

    print(f"📘 Excel 저장 완료: {excel_file}")

    # JSON 저장 (보도시점과 등록일 제거)
    # 보도시점과 등록일은 참고용으로만 사용하므로 저장하지 않음
    results_for_json = []
    for item in results:
        item_copy = {k: v for k, v in item.items() if k not in ['보도시점', '등록일']}
        results_for_json.append(item_copy)
    
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(results_for_json, f, ensure_ascii=False, indent=2)

    print(f"🧾 JSON 저장 완료: {json_file}")


# -----------------------------------------------------------
# 문제가 있는 항목 리스트업 (보도일 없음 또는 보도시점과 1주 이상 차이)
# -----------------------------------------------------------
def list_problematic_items(results):
    """보도일이 없거나 2025년 1월 1일 이전인 항목을 찾습니다
    
    보도일이 2025년 1월 1일 이후면 문제가 아닙니다.
    보도일이 가장 먼저 기준이 됩니다.
    """
    problematic_items = []
    cutoff_date = datetime(2025, 1, 1)
    
    for item in results:
        press_date_str = item.get('보도일')
        
        # 보도일이 없는 경우
        if not press_date_str:
            problematic_items.append({
                **item,
                '문제유형': '보도일 없음'
            })
            continue
        
        # 보도일 파싱
        press_date = parse_date_string(press_date_str)
        
        if not press_date:
            # 보도일 파싱 실패
            problematic_items.append({
                **item,
                '문제유형': '보도일 파싱 실패',
                '원본보도일': press_date_str
            })
            continue
        
        # 보도일이 2025년 1월 1일 이전인 경우만 문제 항목
        if press_date < cutoff_date:
            problematic_items.append({
                **item,
                '문제유형': f'보도일이 2025년 이전 ({press_date.strftime("%Y-%m-%d")})',
                '파싱된보도일': press_date.strftime('%Y-%m-%d')
            })
    
    return problematic_items


# -----------------------------------------------------------
# 실행 메인
# -----------------------------------------------------------
def main():
    import sys
    # 출력 버퍼링 비활성화 (실시간 출력을 위해)
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    
    base_url = "https://www.fss.or.kr/fss/bbs/B0000188/list.do?menuNo=200218&pageIndex=1"
    total_pages = 2010  # 총 페이지 수
    print("금융감독원 보도자료 스크래핑 시작 (2025년 이후 데이터)")
    print("=" * 70)
    sys.stdout.flush()

    results = scrape_press_releases(base_url, total_pages=total_pages, resume=True)

    print("=" * 70)
    print(f"총 {len(results)}개 보도자료 처리 완료")

    success = sum(1 for r in results if r.get('보도일'))
    if results:
        print(f"보도일 추출 성공률: {success}/{len(results)} ({success/len(results)*100:.1f}%)")

    # 최종 저장
    save_results(results)
    
    # 문제가 있는 항목 리스트업
    print("\n" + "=" * 70)
    print("🔍 문제가 있는 항목 분석 중...")
    problematic_items = list_problematic_items(results)
    
    if problematic_items:
        print(f"⚠️ 문제가 있는 항목: {len(problematic_items)}개 발견")
        
        # 문제 항목 저장
        problem_df = pd.DataFrame(problematic_items)
        problem_df['첨부파일'] = problem_df['첨부파일'].apply(
            lambda lst: ', '.join([f"{f['첨부파일명']} ({f['첨부파일 url']})" for f in lst]) if lst else ''
        )
        problem_df.fillna('', inplace=True)
        
        # CSV 저장
        problem_df.to_csv('problematic_items.csv', index=False, encoding='utf-8-sig')
        print(f"📄 문제 항목 CSV 저장 완료: problematic_items.csv")
        
        # Excel 저장
        with pd.ExcelWriter('problematic_items.xlsx', engine='openpyxl') as writer:
            problem_df.to_excel(writer, index=False, sheet_name='문제항목')
            ws = writer.sheets['문제항목']
            for i, col in enumerate(problem_df.columns, start=1):
                max_len = max(problem_df[col].astype(str).map(len).max(), len(col)) + 2
                ws.column_dimensions[get_column_letter(i)].width = min(max_len, 50)
        print(f"📘 문제 항목 Excel 저장 완료: problematic_items.xlsx")
        
        # JSON 저장
        with open('problematic_items.json', 'w', encoding='utf-8') as f:
            json.dump(problematic_items, f, ensure_ascii=False, indent=2)
        print(f"🧾 문제 항목 JSON 저장 완료: problematic_items.json")
        
        # 문제 유형별 통계
        problem_types = {}
        for item in problematic_items:
            ptype = item.get('문제유형', '알 수 없음')
            problem_types[ptype] = problem_types.get(ptype, 0) + 1
        
        print("\n문제 유형별 통계:")
        for ptype, count in problem_types.items():
            print(f"  - {ptype}: {count}개")
    else:
        print("✅ 문제가 있는 항목이 없습니다.")


# -------------------------------------------------
# Health Check 모드
# -------------------------------------------------
from typing import Dict
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import re

def fss_press_releases_health_check() -> Dict:
    """
    금융감독원 보도자료 Health Check
    - 목록 1건 추출
    - 상세 페이지 접근 확인
    """

    BASE_URL = "https://www.fss.or.kr"
    LIST_URL = "https://www.fss.or.kr/fss/bbs/B0000188/list.do?menuNo=200218&pageIndex=1"

    result = {
        "org_name": "FSS",
        "target": "금융감독원 > 보도자료",
        "check_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": "FAIL",
        "checks": {
            "list_page": {},
            "detail_page": {}
        },
        "error": None
    }

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (HealthCheck)"
    })

    try:
        # ===============================
        # 1. 목록 페이지 접근
        # ===============================
        resp = session.get(LIST_URL, timeout=15)
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table", class_="board_list") or soup.find("table")

        if not table:
            result["checks"]["list_page"] = {
                "url": LIST_URL,
                "success": False,
                "count": 0,
                "message": "목록 테이블 없음"
            }
            return result

        rows = table.find_all("tr")[1:]
        if not rows:
            result["checks"]["list_page"] = {
                "url": LIST_URL,
                "success": False,
                "count": 0,
                "message": "목록 데이터 없음"
            }
            return result

        # ===============================
        # 2. 목록 1건 추출
        # ===============================
        first_row = rows[0]
        title_link = first_row.find("a", href=re.compile(r"view\.do"))

        if not title_link:
            result["checks"]["list_page"] = {
                "url": LIST_URL,
                "success": False,
                "count": 0,
                "message": "상세 링크 없음"
            }
            return result

        title = title_link.get_text(strip=True)
        detail_url = urljoin(LIST_URL, title_link["href"])

        result["checks"]["list_page"] = {
            "url": LIST_URL,
            "success": True,
            "count": 1,
            "title": title
        }

        # ===============================
        # 3. 상세 페이지 접근
        # ===============================
        detail_resp = session.get(detail_url, timeout=15)
        detail_resp.raise_for_status()

        detail_soup = BeautifulSoup(detail_resp.text, "html.parser")
        content_div = detail_soup.find("div", class_="dbdata")

        content_length = len(content_div.get_text(strip=True)) if content_div else 0

        result["checks"]["detail_page"] = {
            "url": detail_url,
            "success": True,
            "content_length": content_length
        }

        result["status"] = "OK"
        return result

    except Exception as e:
        result["error"] = str(e)
        result["status"] = "FAIL"
        return result

if __name__ == "__main__":
    import json
    import argparse
    import sys

    parser = argparse.ArgumentParser(description='FSS 보도자료 스크래퍼')
    parser.add_argument("--check", action="store_true", help="FSS 보도자료 Health Check 실행")

    args = parser.parse_args()

    # -------------------------------------------------
    # Health Check 모드
    # python scrape_fss_press_releases_v2.py --check
    # -------------------------------------------------
    if args.check:
        health = fss_press_releases_health_check()
        print(json.dumps(health, ensure_ascii=False, indent=2))
        sys.exit(0)

    main()
