"""
CSV to Excel 변환 유틸리티
"""
import os
import pandas as pd
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def csv_to_excel(
    csv_path: str,
    excel_path: str = None,
    sheet_name: str = "Sheet1",
    column_width_limit: int = 50
) -> str:
    """
    CSV 파일을 Excel 파일로 변환
    
    Args:
        csv_path: CSV 파일 경로
        excel_path: Excel 파일 저장 경로 (None이면 CSV와 같은 디렉토리에 저장)
        sheet_name: Excel 시트 이름
        column_width_limit: 열 너비 최대값
        
    Returns:
        생성된 Excel 파일 경로
    """
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV 파일을 찾을 수 없습니다: {csv_path}")
    
    # Excel 경로 설정
    if excel_path is None:
        csv_dir = os.path.dirname(csv_path)
        csv_name = os.path.splitext(os.path.basename(csv_path))[0]
        excel_dir = os.path.join(csv_dir, "..", "excel")
        excel_dir = os.path.abspath(excel_dir)
        os.makedirs(excel_dir, exist_ok=True)
        excel_path = os.path.join(excel_dir, f"{csv_name}.xlsx")
    
    # Excel 디렉토리 생성
    excel_dir = os.path.dirname(excel_path)
    os.makedirs(excel_dir, exist_ok=True)
    
    # CSV 읽기
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    
    # Excel 저장
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        
        # 열 너비 자동 조정
        for i, col in enumerate(df.columns, start=1):
            max_len = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            ) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(
                max_len, column_width_limit
            )
    
    return excel_path


def data_to_excel(
    data: List[Dict],
    excel_path: str,
    sheet_name: str = "Sheet1",
    column_width_limit: int = 50
) -> str:
    """
    데이터 리스트를 Excel 파일로 저장
    
    Args:
        data: 저장할 데이터 리스트
        excel_path: Excel 파일 저장 경로
        sheet_name: Excel 시트 이름
        column_width_limit: 열 너비 최대값
        
    Returns:
        생성된 Excel 파일 경로
    """
    if not data:
        raise ValueError("저장할 데이터가 없습니다.")
    
    # Excel 디렉토리 생성
    excel_dir = os.path.dirname(excel_path)
    os.makedirs(excel_dir, exist_ok=True)
    
    # DataFrame 생성
    df = pd.DataFrame(data)
    
    # 본문의 줄바꿈 처리 (있는 경우)
    if "본문" in df.columns:
        df["본문"] = df["본문"].astype(str).str.replace("\n", " ").str.replace("\r", " ")
    
    # Excel 저장
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        
        # 열 너비 자동 조정
        for i, col in enumerate(df.columns, start=1):
            max_len = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            ) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(
                max_len, column_width_limit
            )
    
    return excel_path

